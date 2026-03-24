from __future__ import annotations

import csv
import io
import json
from typing import Any

from fastapi import HTTPException


def normalize_records(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        return [dict(item) for item in payload]
    if isinstance(payload, dict):
        if "records" in payload and isinstance(payload["records"], list):
            return [dict(item) for item in payload["records"]]
        return [payload]
    raise HTTPException(status_code=400, detail="Unsupported JSON payload for import.")


def parse_json_bytes(content: bytes) -> list[dict[str, Any]]:
    try:
        payload = json.loads(content.decode("utf-8"))
    except Exception as exc:  # pragma: no cover
        raise HTTPException(status_code=400, detail=f"Invalid JSON import: {exc}") from exc
    return normalize_records(payload)


def parse_csv_bytes(content: bytes) -> list[dict[str, Any]]:
    try:
        decoded = content.decode("utf-8-sig")
    except Exception as exc:  # pragma: no cover
        raise HTTPException(status_code=400, detail=f"Invalid CSV import: {exc}") from exc
    reader = csv.DictReader(io.StringIO(decoded))
    return [dict(row) for row in reader]


def parse_xlsx_bytes(content: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:  # pragma: no cover
        raise HTTPException(
            status_code=500,
            detail="XLSX import requires openpyxl to be installed.",
        ) from exc

    workbook = load_workbook(io.BytesIO(content), read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value) if value is not None else "" for value in rows[0]]
    return [{header: value for header, value in zip(headers, row) if header} for row in rows[1:]]


def parse_upload(filename: str, content: bytes) -> list[dict[str, Any]]:
    lower = filename.lower()
    if lower.endswith(".json"):
        return parse_json_bytes(content)
    if lower.endswith(".csv"):
        return parse_csv_bytes(content)
    if lower.endswith(".xlsx"):
        return parse_xlsx_bytes(content)
    raise HTTPException(status_code=400, detail="Supported import formats: JSON, CSV, XLSX.")

