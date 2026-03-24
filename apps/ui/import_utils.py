from __future__ import annotations

import csv
import io
from datetime import date
from datetime import datetime
from typing import Any


HEADER_ALIASES = {
    "timestamp": {"timestamp", "time", "datetime", "date", "дата", "время", "дата_время"},
    "value": {"value", "measurement", "значение", "измерение"},
    "unit": {"unit", "units", "единица", "единицы"},
    "quality_flag": {"quality_flag", "status", "статус", "качество"},
    "source_type": {"source_type", "source", "источник"},
    "method_reference": {"method_reference", "method", "метод"},
    "accuracy": {"accuracy", "точность"},
    "spatial_location": {"spatial_location", "location", "место", "точка"},
}


def normalize_header(value: Any) -> str:
    return str(value).strip().lower().replace(" ", "_")


def canonical_key(header: str) -> str:
    normalized = normalize_header(header)
    for target, aliases in HEADER_ALIASES.items():
        if normalized in aliases:
            return target
    return normalized


def _rows_from_csv(content: bytes) -> list[dict[str, Any]]:
    decoded = content.decode("utf-8-sig")
    sample = decoded[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(decoded), dialect=dialect)
    return [{canonical_key(key): value for key, value in row.items() if key} for row in reader]


def _rows_from_xlsx(content: bytes) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    all_rows = list(sheet.iter_rows(values_only=True))
    if not all_rows:
        return []

    headers = [canonical_key(item) for item in all_rows[0]]
    return [{header: value for header, value in zip(headers, row) if header} for row in all_rows[1:]]


def parse_measurement_file(filename: str, content: bytes) -> list[dict[str, Any]]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        return _rows_from_csv(content)
    if lower.endswith(".xlsx"):
        return _rows_from_xlsx(content)
    raise ValueError("Поддерживаются только файлы CSV и XLSX.")


def _to_iso(value: Any) -> str:
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.isoformat() + "Z"
        return value.isoformat()
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time()).isoformat() + "Z"
    text = str(value).strip()
    if not text:
        raise ValueError("В файле есть пустая дата или время.")
    return text


def _to_float(value: Any, field_name: str, row_index: int) -> float:
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace(" ", "")
    if not text:
        raise ValueError(f"Строка {row_index}: нет значения для поля '{field_name}'.")

    normalized = text.replace(",", ".")
    try:
        return float(normalized)
    except ValueError as exc:
        raise ValueError(
            f"Строка {row_index}: поле '{field_name}' должно быть числом."
        ) from exc


def prepare_measurement_records(
    rows: list[dict[str, Any]],
    *,
    object_id: str,
    element_id: str,
    channel_id: str,
    default_unit: str,
) -> list[dict[str, Any]]:
    prepared: list[dict[str, Any]] = []
    for index, row in enumerate(rows, start=1):
        if row.get("timestamp") in (None, ""):
            raise ValueError(f"Строка {index}: нет даты/времени.")
        if row.get("value") in (None, ""):
            raise ValueError(f"Строка {index}: нет значения измерения.")

        record = {
            "object_id": object_id,
            "element_id": element_id,
            "channel_id": channel_id,
            "timestamp": _to_iso(row.get("timestamp")),
            "value": _to_float(row.get("value"), "value", index),
            "unit": str(row.get("unit") or default_unit),
        }
        for optional_field in ("quality_flag", "source_type", "method_reference", "accuracy", "spatial_location"):
            value = row.get(optional_field)
            if value not in (None, ""):
                if optional_field == "accuracy":
                    record[optional_field] = _to_float(value, optional_field, index)
                else:
                    record[optional_field] = value
        prepared.append(record)
    return prepared


def preview_rows(rows: list[dict[str, Any]], limit: int = 10) -> list[dict[str, Any]]:
    return rows[:limit]
